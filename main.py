import pandas as pd
from pandas import DataFrame
from pandas.tseries.offsets import BDay
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np
import matplotlib.pyplot as plt
import schedule
import time
import os

def create_finished_excel(sheets: dict) -> None:
     with pd.ExcelWriter('Analysis/Uzduotis_python.xlsx', engine='openpyxl') as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def read_xlsx_file() -> DataFrame:
    df: DataFrame = pd.read_excel('Uzduotis.xlsx', sheet_name=1, engine='openpyxl')
    
    df['Galia VE'] = df['Galia VE'].astype(float)
    df['Galia SE'] = df['Galia SE'].astype(float)
    df['Galia EEKĮ'] = df['Galia EEKĮ'].astype(float)
    df['Dokumentas'] = df['Dokumentas'].astype(str)
    df['Tipas'] = df['Tipas'].astype(str)
    df['1 skyrius'] = df['1 skyrius'].astype(str)
    df['2 skyrius'] = df['2 skyrius'].astype(str)
    df['Gavimo data'] = pd.to_datetime(df['Gavimo data']).dt.date
    df['Atsakymo data'] = pd.to_datetime(df['Atsakymo data']).dt.date
    df['Atsakymo terminas'] = pd.to_datetime(df['Gavimo data'] + pd.offsets.BDay(10)).dt.date  # Add 10 working days
    df['Gavimo data savaitė'] = df['Gavimo data'].apply(lambda x: x.isocalendar()[1] if pd.notnull(x) else np.nan)
    df['Atsakymo data savaitė'] = pd.to_numeric(df['Atsakymo data'].apply(lambda x: x.isocalendar()[1] if pd.notnull(x) else np.nan), errors='coerce').astype('Int64')
    df.fillna(0, inplace=True)   
    df.replace(['None', ''], np.nan, inplace=True)
    return df


def get_ve_and_se_sum(df: DataFrame) -> list:
    filtered_df_ve: DataFrame = df.loc[df['Tipas'].isin(['Hibridas', 'Hibridas+EEKĮ'])]
    filtered_df_se: DataFrame = df.loc[df['Dokumentas'].isin(['Sąlygos', 'Preliminarios sąlygos'])]
    
    return pd.DataFrame({"Galia VE suma": filtered_df_ve['Galia VE'].sum(), "Galia SE suma": round(filtered_df_se['Galia SE'].sum(),2)}, index=[0])


def apply_conditional_formatting(df: DataFrame) -> None:
    try:
        # Load the workbook
        wb = load_workbook('Analysis/Uzduotis_python.xlsx')
        
        # Select the sheet to work with
        ws = wb["Duomenys"]

        # Fill colors for the cells that meet the condition
        late_color = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        on_time_color = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        for idx, row in df.iterrows():
            if row['Atsakymo data'] == 0 or row['Atsakymo data'] <= row['Atsakymo terminas']:
                cell = ws.cell(row=idx+2, column=10)
                cell.fill = on_time_color
            else:
                cell = ws.cell(row=idx+2, column=10)  
                cell.fill = late_color

        wb.save('Analysis/Uzduotis_python.xlsx')
        
    except Exception as e:
        print(f"An error occurred: {e}")
  
        
def plot_document_numbers(df: DataFrame):
    
    def graph_data() -> DataFrame:
        
        # Unique values 
        answered = df['Atsakymo data savaitė'].unique()
        received = df['Gavimo data savaitė'].unique()
        
        graph_df: DataFrame = pd.DataFrame(received, columns=['Gavimo savaitė'])
        graph_df['Gautų dokumentų kiekis'] = df['Gavimo data savaitė'].value_counts()[received].values
        
        graph_df['Atsakymo savaitė'] = answered
        graph_df['Atsakytų dokumentų kiekis'] = df['Atsakymo data savaitė'].value_counts()[answered].values  
        
        return graph_df

    graph_df: DataFrame = graph_data()
    graph_df: DataFrame = graph_df.groupby(['Gavimo savaitė']).sum().reset_index()

    x = range(len(graph_df['Gavimo savaitė']))
    y1 = graph_df['Gautų dokumentų kiekis']
    y2 = graph_df['Atsakytų dokumentų kiekis']
    labels = graph_df['Gavimo savaitė']
    
    plt.figure(figsize=(12, 6))
    plt.bar(x, y1, label='Gautų dokumentų kiekis')
    plt.bar(x, y2, bottom=y1, label='Atsakytų dokumentų kiekis')

    plt.xlabel('Savaitė')
    plt.ylabel('Kiekis')

    plt.xticks(x, labels)
    plt.legend()
    for i, j in zip(x, y1):
        plt.text(i, j, str(j), ha='center', va='top')
    for i, j, k in zip(x, y1, y2):
        plt.text(i, j + k, str(k), ha='center', va='bottom')
    os.makedirs('Analysis', exist_ok=True)
    plt.savefig('Analysis/Grafikas.png', dpi=300)
    

def documents_tool_df(df: DataFrame) -> DataFrame:
    unique_weeks = df['Gavimo data savaitė'].unique()
    tool_df: DataFrame = pd.DataFrame({'Gavimo savaitė': unique_weeks})
    
    counts_for1: list = []
    counts_for2: list = []
    counts_for_both: list = []
    
    for week in unique_weeks:
        
        count_for1 = df[(df['Gavimo data savaitė'] == week) & (df['1 skyrius'] == 'nan')].shape[0]
        count_for2 = df[(df['Gavimo data savaitė'] == week) & (df['2 skyrius'] == 'nan')].shape[0]
        count_for_both = df[(df['Gavimo data savaitė'] == week) & (df['1 skyrius'] == 'nan') & (df['2 skyrius'] == 'nan')].shape[0]
        counts_for1.append(count_for1)
        counts_for2.append(count_for2)
        counts_for_both.append(count_for_both)
    
    tool_df['1 Skyrius'] = counts_for1
    tool_df['2 Skyrius'] = counts_for2
    tool_df["Abu skyriai"] = counts_for_both
    return tool_df


def main() -> None:
    df_data: DataFrame = read_xlsx_file()
    sheets: dict = {'SE ir VE sumos': get_ve_and_se_sum(df_data), 'Dokumentų įrankis': documents_tool_df(df_data), 'Duomenys': df_data}
    plot_document_numbers(df_data)
    create_finished_excel(sheets)
    apply_conditional_formatting(df_data)
    
if __name__ == "__main__":
    main()
    # schedule.every(1).minutes.do(main)
    # while True:
    #     schedule.run_pending()

    
    